import ssl
import re
from urllib.request import urlopen
from bs4 import BeautifulSoup
from openpyxl import Workbook

def getWord():
    context = ssl._create_unverified_context()

    #url 선택부분
    print('''원하는 단어장을 선택하세요 \n 1. 중등\n 2.고등\n 3.토익''')
    selectLevel = int(input("번호 입력 : "))
    print("단어장 만드는 중")

    if selectLevel == 1:
        url ='https://learn.dict.naver.com/m/endic/wordbook/mhs/100001/200001/words.nhn?filterType=0&orderType=2&pageNo='
    elif selectLevel ==2:
        url = 'https://learn.dict.naver.com/m/endic/wordbook/mhs/100004/300059/words.nhn?filterType=0&orderType=2&pageNo='
    elif selectLevel ==3:
        url = 'https://learn.dict.naver.com/m/endic/wordbook/exam/10001/20001/words.nhn?filterType=0&orderType=2&pageNo='
    else:
        print("default")
    
    page = 1

    wb = Workbook()
    ws1 = wb.active
    ws1.title='wordList'
    n=1

    pattern = re.compile(r'\s+')
    for i in range(5):
        newUrl = url+str(page)

        html = urlopen(newUrl,context=context).read()
        soup =BeautifulSoup(html, 'html.parser')
        wordClass=soup.find_all(class_='lst_li2')
        for i in wordClass:
            word = i.select_one('.words').text
            mean = i.select_one('.txt_ct2').text
            splitMean=(mean.split(','))[0]

            splitMean=re.sub(pattern,'',splitMean)

            wordList=[word.splitMean]

            for i in range(n,n+1):
                ws1.cell(row=i, column=1).value=wordList[0]
                ws1.cell(row=i, column=2).value=wordList[1]
            n=n+1
        
        page = page+1
    wb.save('./wordList.xlsx')
